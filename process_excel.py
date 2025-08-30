import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
import shutil

# ダウンロードフォルダに保存
DOWNLOAD_FOLDER = Path.home() / "Downloads"

# 色定義（RGB）
COLORS = [
    "FFF2CC",  # 薄い黄色
    "D9E1F2",  # 薄い青
    "E2EFDA",  # 薄い緑
    "FCE4D6",  # 薄いオレンジ
]

# 手順1: 元VBAのALLマクロの処理
def run_step1(source_file):
    """
    source_file: str または Path - 元Excelファイル
    処理:
    1. 1～3行目を並べ替え
    2. 1～7行目を抽出して4つのExcelに分割
    3. 色付け・列幅設定
    4. ダウンロードフォルダに保存
    """
    source_file = Path(source_file)
    wb = openpyxl.load_workbook(source_file)
    ws = wb.active

    # 並び替え用配列
    sort_order = ["1", "2", "3", "4", ""]

    # 1～3行目を取得して並べ替え
    data = [[ws.cell(row=r, column=c).value for c in range(2, 42)] for r in range(1,4)]
    new_data = [[None]*41 for _ in range(3)]
    k = 0
    for so in sort_order:
        for j in range(41):
            if str(data[0][j]) == so:
                for r in range(3):
                    new_data[r][k] = data[r][j]
                k += 1
    # 書き戻す
    for r in range(3):
        for c in range(41):
            ws.cell(row=r+1, column=c+2, value=new_data[r][c])

    # ファイル名
    file_names = ["調理.xlsx", "接客対応.xlsx", "料理運搬状況確認.xlsx", "宣伝.xlsx"]

    # 1～7行目を4つに分割
    for idx, fname in enumerate(file_names):
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        # A列コピー（1～7行）
        for r in range(1,8):
            new_ws.cell(row=r, column=1, value=ws.cell(row=r, column=1).value)
        col_idx = 2
        for c in range(2, 42):
            if str(ws.cell(row=1, column=c).value) == str(idx+1):
                for r in range(1,8):
                    new_ws.cell(row=r, column=col_idx, value=ws.cell(row=r, column=c).value)
                    # 色付け
                    fill = PatternFill(start_color=COLORS[idx], end_color=COLORS[idx], fill_type="solid")
                    new_ws.cell(row=r, column=col_idx).fill = fill
                col_idx += 1

        # 保存
        save_path = DOWNLOAD_FOLDER / fname
        new_wb.save(save_path)
        print(f"{fname} を {save_path} に保存しました。")

# 手順2: CreateShiftLayoutCleanCentered マクロの処理
def run_step2(source_file):
    """
    source_file: str または Path - 元Excelファイル
    処理:
    1. シートから1～4の列をまとめて新しいシートにコピー
    2. 名前の先頭をファイル名にして保存
    """
    source_file = Path(source_file)
    wb = openpyxl.load_workbook(source_file)
    ws = wb.active

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "ShiftLayout"

    row_max = ws.max_row
    col_max = ws.max_column
    shift_data = []

    # 1～4列目のデータをまとめる
    for col in range(2,5):
        for row in range(1,row_max+1):
            val = ws.cell(row=row, column=col).value
            new_ws.cell(row=row, column=col-1, value=val)

    # ファイル名はA1の値または未分類
    first_name = ws.cell(row=1,column=1).value or "未分類"
    file_name = f"シフト表_{first_name}.xlsx"
    save_path = DOWNLOAD_FOLDER / file_name
    new_wb.save(save_path)
    print(f"{file_name} を {save_path} に保存しました。")
